"""Tests for CARD-030: XLSXParser."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from docforge.config import ParseConfig
from docforge.models import ContentType
from docforge.parsers.xlsx_parser import XLSXParser


def _create_test_xlsx(tmp_path: Path, multi_sheet: bool = True, include_empty: bool = True) -> Path:
    """Create a test XLSX file with sample data."""
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Data"
    ws1["A1"] = "Name"
    ws1["B1"] = "Score"
    ws1["C1"] = "Grade"
    ws1["A2"] = "Alice"
    ws1["B2"] = 95
    ws1["C2"] = "A"
    ws1["A3"] = "Bob"
    ws1["B3"] = 87
    ws1["C3"] = "B"

    if include_empty:
        ws1["A4"] = "Charlie"
        ws1["B4"] = None
        ws1["C4"] = "C"

    if multi_sheet:
        ws2 = wb.create_sheet("Summary")
        ws2["A1"] = "Metric"
        ws2["B1"] = "Value"
        ws2["A2"] = "Average"
        ws2["B2"] = 91
        ws2["A3"] = "Count"
        ws2["B3"] = 3

    filepath = tmp_path / "test.xlsx"
    wb.save(str(filepath))
    wb.close()
    return filepath


class TestXLSXParserCanParse:
    def test_can_parse_xlsx(self):
        parser = XLSXParser()
        assert parser.can_parse(Path("test.xlsx"), "xlsx") is True

    def test_can_parse_other(self):
        parser = XLSXParser()
        assert parser.can_parse(Path("test.pdf"), "pdf") is False
        assert parser.can_parse(Path("test.docx"), "docx") is False


class TestXLSXParserMultiSheet:
    def test_multi_sheet_outputs_markdown_tables(self, tmp_path: Path):
        filepath = _create_test_xlsx(tmp_path, multi_sheet=True)
        config = ParseConfig()
        blocks = XLSXParser().parse(filepath, config)

        table_blocks = [b for b in blocks if b.type == ContentType.TABLE]
        assert len(table_blocks) == 2, (
            f"Expected 2 table blocks (one per sheet), got {len(table_blocks)}"
        )

    def test_sheet_names_as_headings(self, tmp_path: Path):
        filepath = _create_test_xlsx(tmp_path, multi_sheet=True)
        config = ParseConfig()
        blocks = XLSXParser().parse(filepath, config)

        heading_blocks = [b for b in blocks if b.metadata.get("is_heading")]
        assert len(heading_blocks) == 2
        assert "Data" in heading_blocks[0].content
        assert "Summary" in heading_blocks[1].content
        # Check heading format
        assert heading_blocks[0].content.startswith("## Sheet: ")

    def test_single_sheet_works(self, tmp_path: Path):
        filepath = _create_test_xlsx(tmp_path, multi_sheet=False)
        config = ParseConfig()
        blocks = XLSXParser().parse(filepath, config)

        table_blocks = [b for b in blocks if b.type == ContentType.TABLE]
        assert len(table_blocks) == 1


class TestXLSXParserTableContent:
    def test_table_has_correct_structure(self, tmp_path: Path):
        filepath = _create_test_xlsx(tmp_path, multi_sheet=False)
        config = ParseConfig()
        blocks = XLSXParser().parse(filepath, config)

        table_content = blocks[1].content  # Second block should be the table
        assert "|" in table_content
        assert "---" in table_content
        assert "Name" in table_content
        assert "Score" in table_content
        assert "Alice" in table_content
        assert "Bob" in table_content

    def test_empty_cells_handled_gracefully(self, tmp_path: Path):
        filepath = _create_test_xlsx(tmp_path, multi_sheet=False, include_empty=True)
        config = ParseConfig()
        blocks = XLSXParser().parse(filepath, config)

        table_content = blocks[1].content
        assert "Charlie" in table_content
        # Empty cell should not crash; the row should still be present
        lines = table_content.split("\n")
        # At least header + separator + 3 data rows
        assert len(lines) >= 4


class TestXLSXParserEmpty:
    def test_empty_workbook(self, tmp_path: Path):
        wb = Workbook()
        wb.active.title = "Empty"
        filepath = tmp_path / "empty.xlsx"
        wb.save(str(filepath))
        wb.close()

        config = ParseConfig()
        blocks = XLSXParser().parse(filepath, config)
        # Should have a heading block but no table (empty sheet)
        assert len(blocks) == 1
        assert blocks[0].type == ContentType.TEXT
        assert "Empty" in blocks[0].content

    def test_completely_empty_cells_in_row(self, tmp_path: Path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sparse"
        ws["A1"] = "Header"
        ws["B1"] = None
        ws["C1"] = None
        ws["A2"] = "Value"
        filepath = tmp_path / "sparse.xlsx"
        wb.save(str(filepath))
        wb.close()

        config = ParseConfig()
        blocks = XLSXParser().parse(filepath, config)
        assert len(blocks) == 2  # heading + table
        table_content = blocks[1].content
        assert "Header" in table_content
        assert "Value" in table_content
