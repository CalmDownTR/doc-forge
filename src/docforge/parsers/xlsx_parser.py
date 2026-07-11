"""CARD-030: XLSXParser — Excel spreadsheet parser."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from docforge.config import ParseConfig
from docforge.models import ContentBlock, ContentType
from docforge.parsers import BaseParser, register_parser


class XLSXParser(BaseParser):
    """Parser for XLSX (Excel) spreadsheets.

    Uses openpyxl with data_only=True to get formula results.
    Iterates all sheets, outputs sheet name as heading, and converts
    each sheet to a Markdown table.
    """

    def can_parse(self, file_path: Path, file_type: str) -> bool:
        return file_type == "xlsx"

    def parse(self, file_path: Path, config: ParseConfig) -> list[ContentBlock]:
        wb = load_workbook(str(file_path), data_only=True)
        blocks: list[ContentBlock] = []
        reading_order = 0
        page = 1

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Sheet name as heading
            blocks.append(
                ContentBlock(
                    type=ContentType.TEXT,
                    content=f"## Sheet: {sheet_name}",
                    page=page,
                    reading_order=reading_order,
                    metadata={"is_heading": True, "heading_level": 2, "sheet_name": sheet_name},
                )
            )
            reading_order += 1

            # Convert sheet to Markdown table
            markdown_table = self._sheet_to_markdown(ws)
            if markdown_table:
                blocks.append(
                    ContentBlock(
                        type=ContentType.TABLE,
                        content=markdown_table,
                        page=page,
                        reading_order=reading_order,
                        metadata={"sheet_name": sheet_name},
                    )
                )
                reading_order += 1

            page += 1

        wb.close()
        return blocks

    def _sheet_to_markdown(self, ws) -> str:
        """Convert a worksheet to a Markdown table string.

        Empty cells are output as empty strings.
        """
        # Find the used range
        rows_data: list[list[str]] = []
        max_col = 0

        for row in ws.iter_rows():
            row_cells: list[str] = []
            for cell in row:
                value = cell.value
                if value is None:
                    row_cells.append("")
                else:
                    row_cells.append(str(value))
            # Trim trailing empty cells
            while row_cells and row_cells[-1] == "":
                row_cells.pop()
            if row_cells:  # Only include non-empty rows
                rows_data.append(row_cells)
                max_col = max(max_col, len(row_cells))

        if not rows_data:
            return ""

        # Normalize all rows to same column count
        for row in rows_data:
            while len(row) < max_col:
                row.append("")

        lines: list[str] = []
        for row_idx, row_cells in enumerate(rows_data):
            line = "| " + " | ".join(self._escape_cell(cell) for cell in row_cells) + " |"
            lines.append(line)
            if row_idx == 0:
                # Header separator
                sep = "| " + " | ".join("---" for _ in row_cells) + " |"
                lines.append(sep)

        return "\n".join(lines)

    @staticmethod
    def _escape_cell(text: str) -> str:
        """Escape pipe characters in table cell text."""
        return text.replace("|", "\\|")


register_parser("xlsx", XLSXParser)
